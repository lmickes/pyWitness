import openpyxl

class Tester :

    def __init__(self, excelFileName = ""):
        self.excelFileName = excelFileName
        self.excelFile     = openpyxl.load_workbook(self.excelFileName)

    def runAllTests(self, sheet = "Sheet1", outputFileName = "testOutput.xlsx") :
        for i in range(2,self.excelFile[sheet].max_column+1,1) :
            self.runTest(sheet,i)

    def runTest(self, sheet = "Sheet1", column = 2):
        workSheet  = self.excelFile[sheet]
        data       = workSheet.cell(1,column).value
        model      = workSheet.cell(2,column).value
        paramSet   = workSheet.cell(3,column).value
        exclusions = workSheet.cell(4,column).value
        condition  = workSheet.cell(5,column).value
        binning    = workSheet.cell(6,column).value
        paramEstim = workSheet.cell(7,column).value
        maxiter    = workSheet.cell(8,column).value

        exec("dr = __import__('pyWitness')."+data,globals())
        exec("exclusionDict = "+exclusions,globals())

        if binning != "[]" :
            exec("binningArray ="+binning,globals())
            dr.collapseContinuousData(column = "confidence",bins = binningArray,labels=None)

        for k in exclusionDict :
            dr.cutData(k,exclusionDict[k])

        if type(condition) == str :
            conditionColumn = condition.split()[0]
            conditionValue  = condition.split()[1]
            try : 
                conditionValue  = int(condition.split()[1])
            except : 
                pass
            globals()['dp'] = dr.process(conditionColumn,conditionValue)
        else :
            globals()['dp'] = dr.process()

        exec("mf = __import__('pyWitness').ModelFit"+model+"(dp,debug=True)",globals())
        exec("mf.set"+paramSet+"()",globals())

        if paramEstim == "True" :
            mf.setParameterEstimates()

        mf.fit(maxiter=int(maxiter))

        # get results
        mf.printParameters()

        workSheet.cell(10, column, value=mf.lureMean.value)
        workSheet.cell(11, column, value=mf.lureSigma.value)
        workSheet.cell(12, column, value=mf.targetMean.value)
        workSheet.cell(13, column, value=mf.targetSigma.value)
        workSheet.cell(14, column, value=mf.lureBetweenSigma.value)
        workSheet.cell(15, column, value=mf.targetBetweenSigma.value)

        # loop over conditions
        for i in range(0,len(mf.thresholds),1) :
            workSheet.cell(16+i, column, value=getattr(mf,"c"+str(i+1)).value)

        # get other fit numbers
        workSheet.cell(31, column, value=mf.numberIterations)
        workSheet.cell(32, column, value=mf.timeDiff)
        workSheet.cell(33, column, value=mf.chi2)
        workSheet.cell(34, column, value=mf.numberDegreesOfFreedom)
        workSheet.cell(35, column, value=mf.chi2PerNDF)

        # get descriptive stats
        workSheet.cell(37, column, value=dp.numberLineups)
        workSheet.cell(38, column, value=dp.numberTALineups)
        workSheet.cell(39, column, value=dp.numberTPLineups)
        workSheet.cell(40, column, value=dp.data_rates.loc['targetPresent','suspectId'][-1])
        workSheet.cell(41, column, value=dp.data_rates.loc['targetAbsent','suspectId'][-1])
        workSheet.cell(42, column, value=dp.dPrime)
        workSheet.cell(43, column, value=dp.pAUC)

        return globals()['mf']

    def saveWorkbook(self, fileName = "tester.xlsx"):
        self.excelFile.save(fileName)

    def fillResults(self, sheet = "Sheet1", column = 1):
        pass

    def compareResults(self, sheet = "Sheet1", column =1):
        pass



